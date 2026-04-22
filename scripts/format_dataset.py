"""
DatasetFormatterAgent
=====================
A standalone agent that reads an input file of any format (CSV, Excel, JSON,
plain-text, Markdown) and converts it into the spider2-lite JSONL format:

    {"instance_id": "...", "db": "...", "question": "...", "external_knowledge": null}

Supported input formats
-----------------------
* .csv   — each row is a record; the agent maps columns intelligently
* .xlsx / .xls  — same as CSV after reading with openpyxl
* .json  — list of objects or a single object
* .jsonl — one JSON object per line
* .txt / .md — free-form text; an LLM call extracts the questions

Usage
-----
    python scripts/format_dataset.py --input my_questions.csv --db IPL --output data/my_dataset.jsonl

    # With a free-form text file (LLM extraction):
    python scripts/format_dataset.py --input questions.txt --db IPL

    # Override the database for every row:
    python scripts/format_dataset.py --input file.csv --db bank_sales_trading

    # Let the file's own column supply the db:
    python scripts/format_dataset.py --input file.csv
"""
import os
import sys
import re
import json
import csv
import uuid
import argparse
from pathlib import Path
from typing import List, Dict, Optional, Any
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from tt_sql.core.llm_service import LLMService

load_dotenv()

# ─── helpers ──────────────────────────────────────────────────────────────────

def _auto_id(prefix: str, index: int) -> str:
    return f"{prefix}{str(index).zfill(3)}"


def _normalise_columns(row: Dict[str, Any]) -> Dict[str, Any]:
    """Return a lower-cased, stripped copy of a dict's keys."""
    return {k.strip().lower(): v for k, v in row.items()}


_QUESTION_ALIASES   = {"question", "questions", "query", "queries", "q", "text", "prompt", "input", "nl", "natural_language"}
_DB_ALIASES         = {"db", "database", "db_name", "database_name", "table"}
_ID_ALIASES         = {"instance_id", "id", "instance", "idx", "row_id"}
_KB_ALIASES         = {"external_knowledge", "knowledge", "external", "context", "kb"}


def _pick(row: Dict, aliases: set, fallback=None):
    for key in aliases:
        if key in row:
            v = row[key]
            return v if v not in ("", None) else fallback
    return fallback


def _build_record(row: Dict, index: int, default_db: Optional[str],
                  id_prefix: str = "custom") -> Optional[Dict]:
    norm = _normalise_columns(row)
    question  = _pick(norm, _QUESTION_ALIASES)
    if not question:
        return None                 # skip rows with no question
    instance_id = _pick(norm, _ID_ALIASES) or _auto_id(id_prefix, index)
    db          = _pick(norm, _DB_ALIASES)  or default_db or "unknown"
    ext_know    = _pick(norm, _KB_ALIASES)
    return {
        "instance_id": str(instance_id),
        "db": str(db),
        "question": str(question).strip(),
        "external_knowledge": ext_know if ext_know else None,
    }


# ─── readers ──────────────────────────────────────────────────────────────────

class DatasetFormatterAgent:
    """
    Converts any input file into spider2-lite JSONL format.
    For unstructured text files (.txt, .md) an LLM is used to extract questions.
    """

    def __init__(self, llm_service: Optional[LLMService] = None):
        self.llm = llm_service

    # ── readers ──────────────────────────────────────────────────────────────

    def _read_csv(self, path: Path) -> List[Dict]:
        rows = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
        return rows

    def _read_excel(self, path: Path) -> List[Dict]:
        try:
            import openpyxl
        except ImportError:
            import subprocess, sys
            print("[info] openpyxl not found — installing automatically...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for i, v in enumerate(row):
                h = headers[i]
                # skip unnamed/empty columns
                if h and h.lower() not in ("", "none"):
                    rec[h] = v if v is not None else ""
            if rec:
                rows.append(rec)
        return rows

    def _read_json(self, path: Path) -> List[Dict]:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            return [data]
        raise ValueError("JSON file must contain a list or a single object.")

    def _read_jsonl(self, path: Path) -> List[Dict]:
        rows = []
        with open(path, encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    rows.append(json.loads(line))
        return rows

    def _read_text(self, path: Path) -> List[Dict]:
        """Use LLM to extract question objects from free-form text."""
        if not self.llm:
            raise RuntimeError(
                "An LLMService is required to parse unstructured text files. "
                "Make sure your .env includes a valid LLM config."
            )
        text = path.read_text(encoding="utf-8")
        prompt = (
            "You are a data extraction assistant. "
            "Extract all distinct questions or queries from the text below. "
            "Return a JSON array where each element has exactly two keys: "
            '"question" (string) and "external_knowledge" (string or null).\n\n'
            "If a question references a document or file (e.g., 'see doc.md'), "
            'set "external_knowledge" to that filename, otherwise null.\n\n'
            f"TEXT:\n---\n{text[:6000]}\n---\n\n"
            "Return ONLY the JSON array, no explanation."
        )
        raw = self.llm.get_completion(prompt)
        # Strip markdown code fences if present
        raw = re.sub(r"```[a-z]*", "", raw).strip().strip("`")
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            raise ValueError(f"LLM did not return valid JSON. Response was:\n{raw[:400]}")
        return parsed

    # ── orchestrator ──────────────────────────────────────────────────────────

    def convert(self, input_path: str, default_db: Optional[str] = None,
                id_prefix: str = "custom") -> List[Dict]:
        """
        Read *any* file and return a list of spider2-lite records.
        """
        path = Path(input_path)
        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")

        ext = path.suffix.lower()

        if ext == ".csv":
            raw_rows = self._read_csv(path)
        elif ext in (".xlsx", ".xls"):
            raw_rows = self._read_excel(path)
        elif ext == ".json":
            raw_rows = self._read_json(path)
        elif ext == ".jsonl":
            raw_rows = self._read_jsonl(path)
        elif ext in (".txt", ".md"):
            raw_rows = self._read_text(path)
        else:
            # Fallback: try plain text via LLM
            print(f"[warn] Unknown extension '{ext}', treating as plain text.")
            raw_rows = self._read_text(path)

        records = []
        for i, row in enumerate(raw_rows, start=1):
            rec = _build_record(row, i, default_db, id_prefix)
            if rec:
                records.append(rec)

        return records

    def save(self, records: List[Dict], output_path: str) -> None:
        """Write records as JSONL."""
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", encoding="utf-8") as f:
            for rec in records:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        print(f"[ok] Wrote {len(records)} record(s) to {out}")


# ─── CLI entry point ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert any question file into spider2-lite JSONL format."
    )
    parser.add_argument("--input",  "-i", required=True,
                        help="Path to the input file (csv, xlsx, json, jsonl, txt, md).")
    parser.add_argument("--output", "-o", default=None,
                        help="Path for the output JSONL file. "
                             "Defaults to data/<input_stem>.jsonl")
    parser.add_argument("--db", default=None,
                        help="Default database name to use when the input file "
                             "does not supply one.")
    parser.add_argument("--id-prefix", default="custom",
                        help="Prefix for auto-generated instance IDs (default: 'custom').")
    parser.add_argument("--model", default=os.getenv("LLM_MODEL"),
                        help="LLM model to use for unstructured text parsing.")
    args = parser.parse_args()

    # Default output location
    if not args.output:
        stem = Path(args.input).stem
        args.output = str(PROJECT_ROOT / "data" / f"{stem}_formatted.jsonl")

    # Only initialise LLM when we might need it
    ext = Path(args.input).suffix.lower()
    llm = None
    if ext in (".txt", ".md") or ext not in (".csv", ".xlsx", ".xls", ".json", ".jsonl"):
        model = args.model or os.getenv("LLM_MODEL", "gpt-4o")
        print(f"[info] Unstructured input — using LLM ({model}) to extract questions.")
        llm = LLMService(model=model)

    agent = DatasetFormatterAgent(llm_service=llm)

    print(f"[info] Reading: {args.input}")
    records = agent.convert(args.input, default_db=args.db, id_prefix=args.id_prefix)
    print(f"[info] Extracted {len(records)} question(s).")

    agent.save(records, args.output)


if __name__ == "__main__":
    main()
